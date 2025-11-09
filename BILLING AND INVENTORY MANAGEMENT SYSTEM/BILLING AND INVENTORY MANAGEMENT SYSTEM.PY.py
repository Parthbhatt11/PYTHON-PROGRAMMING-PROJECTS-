import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
# --- FIX: Corrected typo 'plat_ypus' to 'platypus' ---
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
# --- NEW: Import for date entry ---
try:
    from tkcalendar import DateEntry
    CALENDAR_ENABLED = True
except ImportError:
    CALENDAR_ENABLED = False
    print("tkcalendar not found. Please install (pip install tkcalendar) for date pickers. Falling back to simple Entry widgets.")


# ------------------- CONFIG -------------------
DATABASE_FILE = "business_app.db"
APP_FONT = ("Segoe UI", 10)
HEADER_FONT = ("Segoe UI", 16, "bold")
DASH_HEADER_FONT = ("Segoe UI", 24, "bold")
DASH_BTN_FONT = ("Segoe UI", 14, "bold")
ACCENT = "#2B7CFF"
ACCENT2 = "#FF6B6B"
# --- NEW: More colors ---
SUCCESS = "#28A745"
PROFIT = "#17A2B8"
BG = "#F2F6FF"
CARD = "#FFFFFF"
TEXT = "#1F2937"
ROW_ODD = "#FFFFFF"
ROW_EVEN = "#F8FBFF"

# ------------------- GLOBAL DATA (IN-MEMORY CACHE) -------------------
bills = []
inventory = {} # Format: {"item_name_lowercase": {"name": "Item Name", "stock": 10, "cost_price": 0, ...}}
business_profile = {"name": "Your Business", "address": "123 Main St", "phone": "555-1234", "gstin": ""}
sale_count = 0
purchase_count = 0
current_items = [] # Temp list for bill form

# --- UI GLOBALS ---
root = None
status_lbl = None
tree = None
items_tree = None
inventory_tree = None
# --- NEW: Report/Customer trees ---
report_tree = None
customer_tree = None
filter_entry = None
type_filter = None
customer_entry, item_entry, qty_entry, price_entry, mode_entry = (None,) * 5
# --- NEW: More dashboard labels ---
lbl_total_sales, lbl_total_purchases, lbl_net, lbl_inventory_value, lbl_total_profit = (None,) * 5
lbl_today_sales, lbl_month_sales = (None,) * 2
items_count_lbl = None
type_var = None
frames = {} # For navigation

# ----------------------------------------------------------------------
# ------------------- PART 1: DATABASE LOGIC (sqlite3) -----------------
# ----------------------------------------------------------------------

def db_connect():
    """Establishes a connection to the SQLite database."""
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Creates/updates the necessary tables."""
    conn = db_connect()
    cursor = conn.cursor()
    
    # Business Profile
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS business_profile (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)
    
    # Inventory Table (Unchanged)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS inventory (
        name_key TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        stock INTEGER NOT NULL DEFAULT 0,
        cost_price REAL NOT NULL DEFAULT 0,
        sale_price REAL NOT NULL DEFAULT 0,
        category TEXT,
        reorder_level INTEGER NOT NULL DEFAULT 5
    )
    """)
    
    # --- DB: Updated Bills Table ---
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS bills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_no INTEGER NOT NULL,
        type TEXT NOT NULL,
        customer TEXT,
        mode TEXT,
        grand_total REAL NOT NULL,
        date TEXT NOT NULL DEFAULT (strftime('%Y-%m-%d','now'))
    )
    """)
    
    # --- DB: Updated Bill Items Table ---
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS bill_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bill_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        qty INTEGER NOT NULL,
        price REAL NOT NULL,
        total REAL NOT NULL,
        cost_price REAL NOT NULL DEFAULT 0,
        FOREIGN KEY (bill_id) REFERENCES bills (id) ON DELETE CASCADE
    )
    """)
    
    # --- DB: Migration helper ---
    def add_column_if_not_exists(table, column, col_type):
        try:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_type}")
        except sqlite3.OperationalError as e:
            if "duplicate column name" not in str(e): raise

    add_column_if_not_exists("inventory", "cost_price", "REAL NOT NULL DEFAULT 0")
    add_column_if_not_exists("inventory", "sale_price", "REAL NOT NULL DEFAULT 0")
    add_column_if_not_exists("inventory", "category", "TEXT")
    add_column_if_not_exists("inventory", "reorder_level", "INTEGER NOT NULL DEFAULT 5")
    # --- NEW: Add date to bills and cost_price to bill_items
    add_column_if_not_exists("bills", "date", "TEXT NOT NULL DEFAULT (strftime('%Y-%m-%d','now'))")
    add_column_if_not_exists("bill_items", "cost_price", "REAL NOT NULL DEFAULT 0")
    
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_bill_type ON bills (bill_no, type)")
    
    conn.commit()
    conn.close()
    
def load_data():
    """Loads all data from SQLite into the global in-memory variables."""
    global bills, inventory, business_profile, sale_count, purchase_count
    
    conn = db_connect()
    cursor = conn.cursor()
    
    # 1. Load Business Profile
    cursor.execute("SELECT key, value FROM business_profile")
    for row in cursor.fetchall():
        if row['key'] in business_profile:
            business_profile[row['key']] = row['value']
            
    # 2. Load Inventory
    inventory = {}
    cursor.execute("SELECT * FROM inventory")
    for row in cursor.fetchall():
        inventory[row['name_key']] = dict(row)
        
    # 3. Load Bills and Bill Items
    bills = []
    cursor.execute("SELECT * FROM bills ORDER BY id")
    all_bills_raw = cursor.fetchall()
    
    for bill_row in all_bills_raw:
        bill = dict(bill_row)
        bill['items'] = []
        cursor.execute("SELECT * FROM bill_items WHERE bill_id = ?", (bill['id'],))
        items_raw = cursor.fetchall()
        for item_row in items_raw:
            bill['items'].append(dict(item_row))
        bills.append(bill)

    # 4. Calculate counts
    sale_count = max((b["bill_no"] for b in bills if b["type"] == "Sale"), default=0)
    purchase_count = max((b["bill_no"] for b in bills if b["type"] == "Purchase"), default=0)
    
    conn.close()
    
    # 5. Update summaries after loading
    update_all_summaries()

def save_business_profile_db():
    conn = db_connect()
    cursor = conn.cursor()
    for key, value in business_profile.items():
        cursor.execute("INSERT OR REPLACE INTO business_profile (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def update_stock_db(item_name, quantity_change):
    """Updates stock in DB and in-memory. Creates item if not exists."""
    global inventory
    key = item_name.lower()
    
    if key not in inventory:
        inventory[key] = {"name": item_name, "stock": 0, "cost_price": 0, "sale_price": 0, "category": None, "reorder_level": 5, "name_key": key}
    
    inventory[key]["stock"] += quantity_change
    new_stock = inventory[key]["stock"]
    
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO inventory (name_key, name, stock, cost_price, sale_price, category, reorder_level)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    ON CONFLICT(name_key) DO UPDATE SET stock = excluded.stock
    """, (key, inventory[key]['name'], new_stock, inventory[key]['cost_price'], inventory[key]['sale_price'], inventory[key]['category'], inventory[key]['reorder_level']))
    conn.commit()
    conn.close()
    
    if inventory_tree:
        refresh_inventory_table() # Refresh UI
    update_main_dashboard_summary() # Update inventory value

def adjust_stock_for_bill(bill_data, action="add"):
    bill_type = bill_data["type"]
    items = bill_data.get("items", [])
    multiplier = 1 if action == "add" else -1
        
    for item in items:
        item_name = item["name"]
        qty = item["qty"]
        if bill_type == "Sale":
            update_stock_db(item_name, -qty * multiplier)
        elif bill_type == "Purchase":
            update_stock_db(item_name, qty * multiplier)

def add_bill_db(bill_data):
    """Adds a new bill and its items to the database."""
    conn = db_connect()
    cursor = conn.cursor()
    
    # --- NEW: Get today's date ---
    today_date = datetime.date.today().strftime('%Y-%m-%d')
    
    try:
        # 1. Insert into main bills table
        cursor.execute("""
        INSERT INTO bills (bill_no, type, customer, mode, grand_total, date)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (bill_data['bill_no'], bill_data['type'], bill_data['customer'], bill_data['mode'], bill_data['grand_total'], today_date))
        
        bill_id = cursor.lastrowid
        
        # 2. Insert all items into bill_items
        items_to_insert = []
        for item in bill_data['items']:
            # --- NEW: Get current cost_price for profit tracking ---
            item_key = item['name'].lower()
            current_cost_price = inventory.get(item_key, {}).get('cost_price', 0)
            
            items_to_insert.append((
                bill_id, item['name'], item['qty'], item['price'], item['total'], current_cost_price
            ))
        
        cursor.executemany("""
        INSERT INTO bill_items (bill_id, name, qty, price, total, cost_price)
        VALUES (?, ?, ?, ?, ?, ?)
        """, items_to_insert)
        
        conn.commit()
        
        # 3. Add to in-memory list
        bill_data['id'] = bill_id
        bill_data['date'] = today_date
        # --- NEW: Add cost_price to in-memory bill items ---
        for i, item in enumerate(bill_data['items']):
            item['cost_price'] = items_to_insert[i][5]  
            
        bills.append(bill_data)
        
        # 4. Adjust stock
        adjust_stock_for_bill(bill_data, action="add")
        
    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Database Error", f"Failed to add bill: {e}")
    finally:
        conn.close()

def edit_bill_db(original_bill, new_bill_data):
    """Updates a bill and its items in the database."""
    conn = db_connect()
    cursor = conn.cursor()
    
    try:
        bill_id = original_bill['id']
        
        # 1. Update the main bill entry
        # Note: We don't update the date of the original bill
        # --- FIX: Use 'bill_no' key ---
        cursor.execute("""
        UPDATE bills SET
            customer = ?, mode = ?, grand_total = ?, type = ?, bill_no = ?
        WHERE id = ?
        """, (new_bill_data['customer'], new_bill_data['mode'], new_bill_data['grand_total'], new_bill_data['type'], new_bill_data['bill_no'], bill_id))
        
        # 2. Delete all old items for this bill
        cursor.execute("DELETE FROM bill_items WHERE bill_id = ?", (bill_id,))
        
        # 3. Insert all new items, with current cost_price
        items_to_insert = []
        new_bill_data['items_with_cost'] = []
        for item in new_bill_data['items']:
            item_key = item['name'].lower()
            current_cost_price = inventory.get(item_key, {}).get('cost_price', 0)
            items_to_insert.append((
                bill_id, item['name'], item['qty'], item['price'], item['total'], current_cost_price
            ))
            # Store for in-memory update
            item_with_cost = dict(item)
            item_with_cost['cost_price'] = current_cost_price
            new_bill_data['items_with_cost'].append(item_with_cost)

        cursor.executemany("""
        INSERT INTO bill_items (bill_id, name, qty, price, total, cost_price)
        VALUES (?, ?, ?, ?, ?, ?)
        """, items_to_insert)
        
        conn.commit()

        # 4. Adjust stock
        adjust_stock_for_bill(original_bill, action="remove") # Revert old
        adjust_stock_for_bill(new_bill_data, action="add")    # Apply new
        
        # 5. Update in-memory list
        for i, b in enumerate(bills):
            if b['id'] == bill_id:
                # Update with new data, but keep original ID and Date
                bills[i].update(new_bill_data)
                bills[i]['items'] = new_bill_data['items_with_cost'] # Update items
                break
                
    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Database Error", f"Failed to update bill: {e}")
    finally:
        conn.close()


def delete_bill_db(bill_to_delete):
    """Deletes a bill and its items from the database."""
    global bills
    conn = db_connect()
    cursor = conn.cursor()
    try:
        bill_id = bill_to_delete['id']
        cursor.execute("DELETE FROM bills WHERE id = ?", (bill_id,)) # Items deleted by CASCADE
        conn.commit()
        adjust_stock_for_bill(bill_to_delete, action="remove")
        bills = [b for b in bills if b['id'] != bill_id]
    except sqlite3.Error as e:
        conn.rollback(); messagebox.showerror("Database Error", f"Failed to delete bill: {e}")
    finally:
        conn.close()

# --- DB: Inventory DB functions (Unchanged) ---
def add_new_product_db(data):
    global inventory
    key = data['name'].lower()
    if key in inventory:
        messagebox.showwarning("Exists", f"Product '{data['name']}' already exists.")
        return False
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("""
        INSERT INTO inventory (name_key, name, stock, cost_price, sale_price, category, reorder_level)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (key, data['name'], data['stock'], data['cost_price'], data['sale_price'], data['category'], data['reorder_level']))
        conn.commit()
        inventory[key] = {"name": data['name'], "stock": data['stock'], "cost_price": data['cost_price'],
                            "sale_price": data['sale_price'], "category": data['category'], "reorder_level": data['reorder_level'], "name_key": key}
        refresh_inventory_table(); update_main_dashboard_summary()
        set_status(f"Added new product: {data['name']}")
        return True
    except sqlite3.Error as e:
        conn.rollback(); messagebox.showerror("Database Error", f"Failed to add product: {e}")
        return False
    finally: conn.close()

def edit_product_db(original_key, data):
    global inventory
    new_key = data['name'].lower()
    if new_key != original_key and new_key in inventory:
        messagebox.showwarning("Exists", f"Product name '{data['name']}' already exists.")
        return False
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("""
        UPDATE inventory SET
            name_key = ?, name = ?, cost_price = ?, sale_price = ?, category = ?, reorder_level = ?
        WHERE name_key = ?
        """, (new_key, data['name'], data['cost_price'], data['sale_price'], data['category'], data['reorder_level'], original_key))
        conn.commit()
        current_stock = inventory[original_key]['stock']
        del inventory[original_key]
        inventory[new_key] = {"name": data['name'], "stock": current_stock, "cost_price": data['cost_price'],
                            "sale_price": data['sale_price'], "category": data['category'],  
                            "reorder_level": data['reorder_level'], "name_key": new_key}
        refresh_inventory_table(); update_main_dashboard_summary()
        set_status(f"Updated product: {data['name']}")
        return True
    except sqlite3.Error as e:
        conn.rollback(); messagebox.showerror("Database Error", f"Failed to update product: {e}")
        return False
    finally: conn.close()

def adjust_product_stock_db(item_key, new_stock):
    global inventory
    if item_key not in inventory: return
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE inventory SET stock = ? WHERE name_key = ?", (new_stock, item_key))
        conn.commit()
        inventory[item_key]["stock"] = new_stock
        refresh_inventory_table(); update_main_dashboard_summary()
        set_status(f"Adjusted stock for {inventory[item_key]['name']}")
    except sqlite3.Error as e:
        conn.rollback(); messagebox.showerror("Database Error", f"Failed to adjust stock: {e}")
    finally: conn.close()

# --- NEW: Delete Product DB Function ---
def delete_product_db(item_key):
    global inventory
    if item_key not in inventory:
        messagebox.showerror("Error", "Could not find product to delete.")
        return
    
    item_name = inventory[item_key]['name']

    # Check if item is used in any bill
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM bill_items WHERE name = ? LIMIT 1", (item_name,))
    exists = cursor.fetchone()
    
    if exists:
        if not messagebox.askyesno("Confirm Delete", f"Product '{item_name}' is present in past bills.\n\nDeleting it may affect historical data if you regenerate reports (though profit reports should be fine).\n\nAre you sure you want to permanently delete this product?"):
            conn.close()
            return

    try:
        # Proceed with deletion
        cursor.execute("DELETE FROM inventory WHERE name_key = ?", (item_key,))
        conn.commit()
        
        # Remove from in-memory cache
        del inventory[item_key]
        
        refresh_inventory_table()
        update_main_dashboard_summary()
        set_status(f"Deleted product: {item_name}")
        messagebox.showinfo("Deleted", f"Product '{item_name}' has been deleted.")

    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Database Error", f"Failed to delete product: {e}")
    finally:
        conn.close()


# --- DB: Query Functions ---
def get_inventory_value():
    # This query can be slow if inventory is huge, but for SQLite it's fine.
    # An alternative is to cache this value.
    if not inventory: return 0
    return sum(item.get('stock', 0) * item.get('cost_price', 0) for item in inventory.values())

def get_total_profit():
    # Calculates profit from all *PAST* sales.
    conn = db_connect()
    cursor = conn.cursor()
    try:
        # Profit = (Sale Price - Cost Price) * Quantity
        # We stored total = Sale Price * Qty
        # We stored cost_price = Cost Price
        # So profit = total - (cost_price * qty)
        cursor.execute("""
        SELECT SUM(total) - SUM(cost_price * qty) 
        FROM bill_items 
        WHERE bill_id IN (SELECT id FROM bills WHERE type = 'Sale')
        """)
        result = cursor.fetchone()[0]
        return result if result else 0
    except Exception:
        return 0
    finally:
        conn.close()

def get_sales_for_period(start_date, end_date):
    """Gets total sales amount for a given period."""
    conn = db_connect()
    cursor = conn.cursor()
    try:
        cursor.execute("""
        SELECT SUM(grand_total) 
        FROM bills 
        WHERE type = 'Sale' AND date BETWEEN ? AND ?
        """, (start_date, end_date))
        result = cursor.fetchone()[0]
        return result if result else 0
    except Exception:
        return 0
    finally:
        conn.close()

# ----------------------------------------------------------------------
# ------------------- PART 2: CORE APP LOGIC ---------------------------
# ----------------------------------------------------------------------

# --- NEW: Changed to "Rs." ---
def format_currency(x):
    if x is None: x = 0
    return f"Rs. {x:,.2f}"

def set_status(text, timeout=3000):
    if status_lbl:
        status_lbl.config(text=text)
        if timeout:
            root.after(timeout, lambda: status_lbl.config(text="Ready"))

def update_all_summaries():
    """Calls all summary update functions."""
    update_billing_summary()
    update_main_dashboard_summary()

def update_billing_summary():
    """Updates the billing dashboard cards."""
    total_sales = sum(b.get("grand_total", 0) for b in bills if b["type"] == "Sale")
    total_purchases = sum(b.get("grand_total", 0) for b in bills if b["type"] == "Purchase")
    net = total_sales - total_purchases
    
    # --- NEW: Get sales for today and this month ---
    today = datetime.date.today().strftime('%Y-%m-%d')
    month_start = datetime.date.today().strftime('%Y-%m-01')
    
    today_sales = get_sales_for_period(today, today)
    month_sales = get_sales_for_period(month_start, today)
    
    if lbl_total_sales: lbl_total_sales.config(text=format_currency(total_sales))
    if lbl_total_purchases: lbl_total_purchases.config(text=format_currency(total_purchases))
    if lbl_net:
        lbl_net.config(text=format_currency(net))
        lbl_net.config(fg="#0A7A0A" if net >= 0 else "#B00020")
    # --- NEW: Update today/month labels ---
    if lbl_today_sales: lbl_today_sales.config(text=format_currency(today_sales))
    if lbl_month_sales: lbl_month_sales.config(text=format_currency(month_sales))

def update_main_dashboard_summary():
    """Updates the main dashboard cards."""
    if lbl_inventory_value:
        lbl_inventory_value.config(text=format_currency(get_inventory_value()))
    # --- NEW: Update profit label ---
    if lbl_total_profit:
        lbl_total_profit.config(text=format_currency(get_total_profit()))

def clear_entries():
    if customer_entry: customer_entry.delete(0, tk.END)
    if item_entry: item_entry.delete(0, tk.END)
    if qty_entry: qty_entry.delete(0, tk.END)
    if price_entry: price_entry.delete(0, tk.END)
    if mode_entry: mode_entry.delete(0, tk.END)
    if type_var: type_var.set("Sale")
    clear_current_items()

def validate_integer(P):
    if P == "" or P == "-": return True
    try: int(P); return True
    except ValueError: return False

def validate_float(P):
    if P == "" or P == "-" or P == ".": return True
    if P.count('.') > 1: return False
    try: float(P); return True
    except ValueError: return False

def edit_business_profile():
    global business_profile
    # This still uses simpledialog, which is fine for a rarely-used feature.
    name = simpledialog.askstring("Business Name", "Enter Business Name:", initialvalue=business_profile["name"])
    if name is None: return
    addr = simpledialog.askstring("Address", "Enter Business Address:", initialvalue=business_profile["address"])
    if addr is None: return
    phone = simpledialog.askstring("Phone", "Enter Business Phone:", initialvalue=business_profile["phone"])
    if phone is None: return
    gstin = simpledialog.askstring("GSTIN", "Enter GSTIN (if any):", initialvalue=business_profile["gstin"])
    if gstin is None: return
    
    business_profile.update({"name": name, "address": addr, "phone": phone, "gstin": gstin})
    save_business_profile_db()
    messagebox.showinfo("Saved", "✅ Business details updated successfully.")
    set_status("Business details updated")

# --- ITEM MANAGEMENT (UI) ---
def refresh_items_tree():
    if not items_tree: return
    for i in items_tree.get_children():
        items_tree.delete(i)
    for idx, it in enumerate(current_items, start=1):
        items_tree.insert("", tk.END, values=(idx, it["name"], it["qty"], format_currency(it["price"]), format_currency(it["total"])))
    if items_count_lbl:
        items_count_lbl.config(text=f"Items: {len(current_items)}")

def get_stock(item_name):
    key = item_name.lower()
    return inventory.get(key, {}).get("stock", 0)

def check_stock_availability(items_list):
    for item in items_list:
        item_name = item["name"]
        qty_needed = item["qty"]
        stock_available = get_stock(item_name)
        if qty_needed > stock_available:
            return f"Not enough stock for '{item_name}'.\nNeed: {qty_needed}, Available: {stock_available}"
    return True

def auto_fill_item_price():
    """Auto-fills price when item name is entered."""
    key = item_entry.get().lower()
    if key in inventory:
        price_entry.delete(0, tk.END)
        price_entry.insert(0, str(inventory[key]['sale_price']))
    
def add_item_to_current():
    name = item_entry.get().strip()
    qty_text = qty_entry.get().strip()
    price_text = price_entry.get().strip()
    if not name:
        messagebox.showerror("Invalid Input", "Item name cannot be empty."); return
    try: qty = int(qty_text)
    except Exception:
        messagebox.showerror("Invalid Input", "Quantity must be a positive integer."); return
    try: price = float(price_text)
    except Exception:
        messagebox.showerror("Invalid Input", "Price must be a non-negative number."); return
        
    if type_var.get() == "Sale":
        qty_needed = qty + sum(it["qty"] for it in current_items if it["name"].lower() == name.lower())
        stock_available = get_stock(name)
        if qty_needed > stock_available:
            messagebox.showwarning("Low Stock", 
                                   f"Warning: Not enough stock for '{name}'.\n"
                                   f"Total Needed: {qty_needed}, Available: {stock_available}\n"
                                   "You can proceed, but stock will go negative.")
            
    current_items.append({"name": name, "qty": qty, "price": price, "total": qty * price})
    refresh_items_tree()
    item_entry.delete(0, tk.END); qty_entry.delete(0, tk.END); price_entry.delete(0, tk.END)
    item_entry.focus()
    set_status("Item added to current bill")

def remove_selected_item():
    if not items_tree: return
    sel = items_tree.selection()
    if not sel: return
    idx = items_tree.index(sel[0])
    del current_items[idx]
    refresh_items_tree(); set_status("Removed item")

def clear_current_items():
    current_items.clear(); refresh_items_tree()

# --- MAIN TABLE ---
def first_item_display(b):
    items = b.get("items")
    if items:
        return f"{items[0]['name']} (+{len(items)-1} more)" if len(items) > 1 else items[0]["name"]
    return ""

def refresh_table(filter_text="", filter_type="All"):
    if not tree: return
    for i in tree.get_children():
        tree.delete(i)
    
    filtered = bills
    if filter_type in ("Sale", "Purchase"):
        filtered = [b for b in filtered if b["type"] == filter_type]
    if filter_text:
        ft = filter_text.lower()
        filtered = [b for b in filtered if (ft in str(b["bill_no"]).lower()
                        or ft in b.get("customer", "").lower()
                        or ft in first_item_display(b).lower())]
                        
    for idx, b in enumerate(filtered, start=1):
        tag = "even" if idx % 2 == 0 else "odd"
        tree.insert("", tk.END, values=(
            idx, b["bill_no"], b["type"], b.get("customer", ""),
            first_item_display(b),
            sum(it.get("qty", 0) for it in b.get("items", [])),
            format_currency(b.get("grand_total", 0)),
            b.get("mode", ""),
            b.get("date", "") # --- NEW: Show date
        ), tags=(tag,), iid=b['id'])
    
    update_billing_summary()

# --- FORM DATA ---
def get_form_data():
    customer = customer_entry.get().strip()
    mode = mode_entry.get().strip()
    typ = type_var.get()
    
    if not customer:
        messagebox.showerror("Invalid Input", "Customer/Supplier cannot be empty."); return None
        
    items = []
    if current_items:
        items = [dict(it) for it in current_items]
    else: 
        item = item_entry.get().strip()
        if not item:
            messagebox.showerror("Invalid Input", "Item cannot be empty."); return None
        try: qty = int(qty_entry.get().strip())
        except Exception:
            messagebox.showerror("Invalid Input", "Quantity must be a positive integer."); return None
        try: price = float(price_entry.get().strip())
        except Exception:
            messagebox.showerror("Invalid Input", "Price must be a non-negative number."); return None
        items = [{"name": item, "qty": qty, "price": price, "total": qty * price}]
        
    if not items:
         messagebox.showerror("Invalid Input", "Bill must contain at least one item."); return None
         
    if typ == "Sale":
        stock_check = check_stock_availability(items)
        if stock_check is not True:
            if not messagebox.askyesno("Low Stock", f"{stock_check}\n\nProceed anyway?"):
                return None
            
    if mode == "":
        messagebox.showerror("Invalid Input", "Mode cannot be empty (Cash/Credit)."); return None
        
    grand_total = sum(it["total"] for it in items)
    return {"type": typ, "customer": customer, "items": items, "mode": mode, "grand_total": grand_total}

# --- CRUD (Create, Read, Update, Delete) ---

# --- FIX: Fixed 'billNo' vs 'bill_no' and reset filters ---
def add_bill():
    global sale_count, purchase_count
    data = get_form_data()
    if not data: return
    
    if data["type"] == "Sale":
        sale_count += 1; bill_no = sale_count
    else:
        purchase_count += 1; bill_no = purchase_count
        
    bill_data = {"bill_no": bill_no, **data} # --- FIX: Use 'bill_no'
    
    add_bill_db(bill_data)
    
    # --- FIX: Reset filters to ensure new bill is visible ---
    if filter_entry: filter_entry.delete(0, tk.END)
    if type_filter: type_filter.set("All")
    refresh_table("", "All") 
    # --- END FIX ---
    
    clear_entries()
    set_status(f"Added {data['type']} Bill #{bill_no}")
    messagebox.showinfo("Added", f"✅ {data['type']} Bill #{bill_no} added.")
    update_all_summaries()

def on_row_select(event):
    if not tree: return
    sel = tree.focus()
    if not sel: return
    
    bill_id = int(sel)
    bill = next((b for b in bills if b["id"] == bill_id), None)
    
    if not bill:
        clear_entries(); return

    clear_entries()
    type_var.set(bill["type"])
    customer_entry.insert(0, bill["customer"])
    
    clear_current_items()
    for it in bill.get("items", []):
        current_items.append(dict(it))
    refresh_items_tree()
    
    if current_items:
        it0 = current_items[0]
        item_entry.insert(0, it0["name"])
        qty_entry.insert(0, str(it0["qty"]))
        price_entry.insert(0, str(it0["price"]))
    mode_entry.insert(0, bill.get("mode", ""))

# --- FIX: Fixed 'billNo' vs 'bill_no' and reset filters ---
def edit_bill():
    if not tree: return
    sel = tree.focus()
    if not sel:
        messagebox.showwarning("Select", "Select a bill to edit."); return
        
    bill_id = int(sel)
    original_bill = next((b for b in bills if b["id"] == bill_id), None)
    if not original_bill:
        messagebox.showerror("Not Found", "Bill not found in memory."); return

    new_data = get_form_data()
    if not new_data: return
    
    # --- FIX: Use 'bill_no' ---
    if new_data['type'] == original_bill['type']:
        new_data['bill_no'] = original_bill['bill_no']
    else:
        global sale_count, purchase_count
        if new_data["type"] == "Sale":
            sale_count += 1; new_data['bill_no'] = sale_count
        else:
            purchase_count += 1; new_data['bill_no'] = purchase_count
    
    edit_bill_db(original_bill, new_data)
    
    # --- FIX: Reset filters ---
    if filter_entry: filter_entry.delete(0, tk.END)
    if type_filter: type_filter.set("All")
    refresh_table("", "All")
    # --- END FIX ---
    
    clear_entries()
    set_status(f"Updated Bill #{original_bill['bill_no']}")
    messagebox.showinfo("Updated", f"✏️ Bill #{original_bill['bill_no']} updated.")
    update_all_summaries()

# --- FIX: Reset filters on delete ---
def delete_bill():
    if not tree: return
    sel = tree.focus()
    if not sel:
        messagebox.showwarning("Select", "Select a bill to delete."); return
        
    bill_id = int(sel)
    bill_to_delete = next((b for b in bills if b["id"] == bill_id), None)
    if not bill_to_delete:
        messagebox.showerror("Not Found", "Bill not found in memory."); return
        
    bill_no, bill_type = bill_to_delete['bill_no'], bill_to_delete['type']
    
    if not messagebox.askyesno("Confirm", f"Delete {bill_type} Bill #{bill_no}?"): return
    
    delete_bill_db(bill_to_delete)
    
    # --- FIX: Reset filters ---
    if filter_entry: filter_entry.delete(0, tk.END)
    if type_filter: type_filter.set("All")
    refresh_table("", "All")
    # --- END FIX ---
    
    clear_entries()
    set_status(f"Deleted Bill #{bill_no}")
    messagebox.showinfo("Deleted", f"🗑️ Bill #{bill_no} deleted.")
    update_all_summaries()

# --- PDF / EXPORT (Unchanged) ---
# --- NEW: Revamped PDF creation for better design ---
def create_invoice_pdf():
    if not tree:
        return
    sel = tree.focus()
    if not sel:
        messagebox.showwarning("Select Bill", "Select a bill to create invoice.")
        return

    bill_id = int(sel)
    bill_data = next((b for b in bills if b["id"] == bill_id), None)
    if not bill_data:
        messagebox.showerror("Not Found", "Bill data not found.")
        return

    fpath = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")],
        initialfile=f"Invoice_{bill_data['type']}_{bill_data['bill_no']}.pdf",
        title="Save Invoice PDF"
    )
    if not fpath:
        return

    # --- Font setup ---
    try:
        font_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DejaVuSans.ttf")
        if not os.path.exists(font_path):
            font_path = "DejaVuSans.ttf"
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        bold_font_path = font_path.replace("Sans.ttf", "Sans-Bold.ttf")
        if os.path.exists(bold_font_path):
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', bold_font_path))
        active_font = 'DejaVuSans'
        active_font_bold = 'DejaVuSans-Bold'
    except Exception:
        active_font = 'Helvetica'
        active_font_bold = 'Helvetica-Bold'

    # --- PDF setup ---
    doc = SimpleDocTemplate(
        fpath, pagesize=A4,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch
    )
    styles = getSampleStyleSheet()

    # --- Custom styles (safe add) ---
    def add_style_safe(name, **kwargs):
        if name not in styles:
            styles.add(ParagraphStyle(name=name, **kwargs))
        else:
            for k, v in kwargs.items():
                setattr(styles[name], k, v)

    add_style_safe('CustomTitle', fontName=active_font_bold, fontSize=20,
                   alignment=0, textColor=colors.HexColor(ACCENT))
    add_style_safe('BusinessInfo', fontName=active_font, fontSize=10, alignment=0)
    add_style_safe('InvoiceHeader', fontName=active_font_bold, fontSize=12, alignment=2)
    add_style_safe('BillTo', fontName=active_font, fontSize=10, alignment=0)
    add_style_safe('TotalText', fontName=active_font_bold, fontSize=12, alignment=2)
    add_style_safe('TotalAmount', fontName=active_font_bold, fontSize=12,
                   alignment=2, textColor=colors.HexColor(ACCENT))
    add_style_safe('Footer', fontName=active_font, fontSize=9, alignment=1)

    elements = []

    # --- Header ---
    header_data = [
        [Paragraph(business_profile['name'], styles['CustomTitle']),
         Paragraph(f"INVOICE / {bill_data['type'].upper()}", styles['InvoiceHeader'])],
        [Paragraph(business_profile['address'], styles['BusinessInfo']),
         Paragraph(f"<b>Bill No:</b> {bill_data['bill_no']}", styles['BusinessInfo'])],
        [Paragraph(f"Phone: {business_profile['phone']}", styles['BusinessInfo']),
         Paragraph(f"<b>Date:</b> {bill_data.get('date', 'N/A')}", styles['BusinessInfo'])],
        [Paragraph(f"GSTIN: {business_profile['gstin']}", styles['BusinessInfo']),
         Paragraph(f"<b>Mode:</b> {bill_data['mode']}", styles['BusinessInfo'])]
    ]
    header_table = Table(header_data, colWidths=[3.5 * inch, 3.5 * inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT')
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.25 * inch))

    # --- Bill To section ---
    bill_to_data = [
        [Paragraph("<b>BILL TO:</b>", styles['BillTo'])],
        [Paragraph(bill_data['customer'], styles['BillTo'])]
    ]
    bill_to_table = Table(bill_to_data, colWidths=[doc.width])
    bill_to_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey)
    ]))
    elements.append(bill_to_table)
    elements.append(Spacer(1, 0.25 * inch))

    # --- Items Table ---
    data = [["S.No", "Item Description", "Qty", "Price (Rs.)", "Total (Rs.)"]]
    for i, it in enumerate(bill_data.get("items", []), start=1):
        data.append([
            str(i),
            Paragraph(it["name"], styles['BodyText']),
            str(it["qty"]),
            format_currency(it["price"]),
            format_currency(it["total"])
        ])

    table = Table(data, colWidths=[0.5 * inch, 3.5 * inch, 0.75 * inch, 1.15 * inch, 1.15 * inch])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), active_font_bold),
        ('FONTNAME', (0, 1), (-1, -1), active_font),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(ACCENT)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6)
    ]))
    elements.append(table)

    # --- Totals ---
    total_data = [
        [Paragraph("Grand Total:", styles['TotalText']),
         Paragraph(format_currency(bill_data.get("grand_total", 0)), styles['TotalAmount'])]
    ]
    total_table = Table(total_data, colWidths=[5.5 * inch, 1.5 * inch])
    total_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT')
    ]))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(total_table)

    # --- Footer ---
    elements.append(Spacer(1, 0.5 * inch))
    elements.append(Paragraph("THANK YOU FOR YOUR BUSINESS!", styles['Footer']))

    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(active_font, 9)
        canvas.drawCentredString(A4[0] / 2, 0.25 * inch,
                                f"Page {doc.page} | {business_profile['name']}")
        canvas.restoreState()

    # --- Build PDF ---
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    messagebox.showinfo("Invoice Created", f"✅ Invoice PDF saved as {os.path.basename(fpath)}")
    set_status(f"Invoice PDF created for Bill #{bill_data['bill_no']}")


def show_ledger():
    if not customer_entry: return
    name = customer_entry.get().strip()
    if not name:
        messagebox.showerror("Input required", "Enter Customer/Supplier name to view ledger."); return
    total_sales = sum(b.get("grand_total", 0) for b in bills if b["customer"].lower() == name.lower() and b["type"] == "Sale")
    total_purchases = sum(b.get("grand_total", 0) for b in bills if b["customer"].lower() == name.lower() and b["type"] == "Purchase")
    balance = total_sales - total_purchases
    messagebox.showinfo("Ledger Summary",
                        f"Name: {name}\n\nTotal Sales: {format_currency(total_sales)}\n"
                        f"Total Purchases: {format_currency(total_purchases)}\n\n"
                        f"Balance: {format_currency(balance)}")

# --- NEW: Renamed to export_bills_excel ---
def export_bills_excel():
    if not bills:
        messagebox.showwarning("No Data", "No bills to export."); return
    fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],
                                         initialfile="bills_data.xlsx", title="Export bills to Excel")
    if not fpath: return
    wb = Workbook(); ws = wb.active; ws.title = "Bills"
    headers = ["S.No", "BillNo", "Date", "Type", "Customer", "Items (name x qty)", "QtyTotal", "PriceSummary", "Mode", "Grand Total"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
        
    for idx, b in enumerate(bills, start=1):
        items_text = "; ".join(f"{it['name']} x{it['qty']}" for it in b.get("items", []))
        qty_total = sum(it.get("qty", 0) for it in b.get("items", []))
        price_summary = "; ".join(format_currency(it['price']) for it in b.get("items", []))
        ws.append([
            idx, b["bill_no"], b.get('date', ''), b["type"], b.get("customer", ""), items_text, qty_total, price_summary, b.get("mode", ""), b.get("grand_total", 0)
        ])

    auto_size_excel_columns(ws)
    
    wb.save(fpath)
    set_status(f"Exported {len(bills)} bills to {os.path.basename(fpath)}")
    messagebox.showinfo("Exported", f"✅ Exported {len(bills)} bills to {os.path.basename(fpath)}")

# --- NEW: Helper function to auto-size columns in Excel ---
def auto_size_excel_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

# --- NEW: Export Inventory ---
def export_inventory_excel():
    if not inventory:
        messagebox.showwarning("No Data", "No inventory items to export."); return
    fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],
                                        initialfile="inventory_data.xlsx", title="Export Inventory to Excel")
    if not fpath: return
    
    wb = Workbook(); ws = wb.active; ws.title = "Inventory"
    headers = ["S.No", "Product Name", "Category", "Stock", "Reorder Lvl", "Cost Price", "Sale Price"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

    sorted_items = sorted(list(inventory.values()), key=lambda x: x["name"])
    for idx, item in enumerate(sorted_items, start=1):
        ws.append([
            idx, item["name"], item.get("category", "N/A"), item.get("stock", 0),
            item.get("reorder_level", 0), item.get("cost_price", 0), item.get("sale_price", 0)
        ])
    
    auto_size_excel_columns(ws)
    wb.save(fpath)
    set_status(f"Exported {len(sorted_items)} inventory items")
    messagebox.showinfo("Exported", f"✅ Exported {len(sorted_items)} inventory items to {os.path.basename(fpath)}")

# --- NEW: Export Report ---
def export_report_excel():
    if not report_tree or not report_tree.get_children():
        messagebox.showwarning("No Data", "No report data to export. Run a report first."); return
    fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],
                                         initialfile="sales_report.xlsx", title="Export Report to Excel")
    if not fpath: return

    wb = Workbook(); ws = wb.active; ws.title = "Sales Report"
    headers = ["Item Name", "Units Sold", "Total Revenue", "Total Cost", "Total Profit"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

    for item_id in report_tree.get_children():
        values = report_tree.item(item_id)['values']
        ws.append(values)
    
    auto_size_excel_columns(ws)
    wb.save(fpath)
    set_status("Exported sales report")
    messagebox.showinfo("Exported", f"✅ Exported sales report to {os.path.basename(fpath)}")

# --- NEW: Export Customers ---
def export_customers_excel():
    if not customer_tree or not customer_tree.get_children():
        messagebox.showwarning("No Data", "No customer data to export. Refresh the list first."); return
    fpath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],
                                         initialfile="customer_list.xlsx", title="Export Customers to Excel")
    if not fpath: return

    wb = Workbook(); ws = wb.active; ws.title = "Customers"
    headers = ["Customer Name", "Total Bills", "Total Spent"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

    for item_id in customer_tree.get_children():
        values = customer_tree.item(item_id)['values']
        ws.append(values)
    
    auto_size_excel_columns(ws)
    wb.save(fpath)
    set_status("Exported customer list")
    messagebox.showinfo("Exported", f"✅ Exported customer list to {os.path.basename(fpath)}")


# --- SEARCH / FILTER ---
def on_search_change(*_): 
    if filter_entry and type_filter: refresh_table(filter_entry.get(), type_filter.get())
def on_filter_change(*_): 
    if filter_entry and type_filter: refresh_table(filter_entry.get(), type_filter.get())

# ----------------------------------------------------------------------
# ------------------- PART 3: INVENTORY UI LOGIC -----------------------
# ----------------------------------------------------------------------

# --- UI: New Custom Dialog for Add/Edit Product ---
class ProductEditDialog(tk.Toplevel):
    def __init__(self, parent, product_data=None):
        super().__init__(parent)
        self.transient(parent); self.grab_set()
        self.parent = parent; self.product_data = product_data; self.result = None
        self.configure(bg=CARD)
        self.is_edit_mode = product_data is not None
        self.title(f"{'Edit' if self.is_edit_mode else 'Add'} Product")
        
        self.main_frame = ttk.Frame(self, padding=20, style="Card.TFrame")
        self.main_frame.pack(fill="both", expand=True)
        
        self.vcmd_int = (self.register(validate_integer), '%P')
        self.vcmd_float = (self.register(validate_float), '%P')

        self.entries = {}
        fields = ["Name", "Category", "Cost Price", "Sale Price", "Reorder Level"]
        if not self.is_edit_mode:
            fields.insert(2, "Initial Stock")
        
        for i, field in enumerate(fields):
            lbl = ttk.Label(self.main_frame, text=f"{field}:", style="TLabel", background=CARD)
            lbl.grid(row=i, column=0, sticky="w", padx=10, pady=5)
            entry = ttk.Entry(self.main_frame, width=40)
            if field in ["Cost Price", "Sale Price"]: entry.config(validate="key", validatecommand=self.vcmd_float)
            elif field in ["Initial Stock", "Reorder Level"]: entry.config(validate="key", validatecommand=self.vcmd_int)
            entry.grid(row=i, column=1, sticky="w", padx=10, pady=5)
            self.entries[field] = entry

        if self.is_edit_mode:
            self.entries["Name"].insert(0, product_data.get("name", ""))
            self.entries["Category"].insert(0, product_data.get("category", "") or "")
            self.entries["Cost Price"].insert(0, product_data.get("cost_price", 0))
            self.entries["Sale Price"].insert(0, product_data.get("sale_price", 0))
            self.entries["Reorder Level"].insert(0, product_data.get("reorder_level", 5))
            
        btn_frame = ttk.Frame(self.main_frame, style="Card.TFrame")
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=20)
        ttk.Button(btn_frame, text="Save", command=self.on_ok, style="Accent.TButton").pack(side="left", padx=10)
        ttk.Button(btn_frame, text="Cancel", command=self.destroy, style="TButton").pack(side="left", padx=10)
        
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        self.wait_window()

    def on_ok(self):
        try:
            data = {'name': self.entries["Name"].get().strip()}
            if not data['name']:
                messagebox.showerror("Error", "Name is required.", parent=self); return
            data['category'] = self.entries["Category"].get().strip() or None
            data['cost_price'] = float(self.entries["Cost Price"].get() or 0)
            data['sale_price'] = float(self.entries["Sale Price"].get() or 0)
            data['reorder_level'] = int(self.entries["Reorder Level"].get() or 5)
            
            if self.is_edit_mode:
                data['stock'] = self.product_data['stock']
                success = edit_product_db(self.product_data['name_key'], data)
            else:
                data['stock'] = int(self.entries["Initial Stock"].get() or 0)
                success = add_new_product_db(data)
            if success: self.destroy()
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid input for numeric field: {e}", parent=self)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}", parent=self)

def refresh_inventory_table(low_stock_only=False):
    """Updates the inventory Treeview UI, with optional low-stock filter."""
    if not inventory_tree: return
    for i in inventory_tree.get_children():
        inventory_tree.delete(i)
    
    items_to_display = []
    if low_stock_only:
        items_to_display = [item for item in inventory.values() if item['stock'] <= item['reorder_level']]
    else:
        items_to_display = list(inventory.values())
    sorted_items = sorted(items_to_display, key=lambda x: x["name"])
    
    for idx, item in enumerate(sorted_items, start=1):
        tag = "even" if idx % 2 == 0 else "odd"
        stock = item.get("stock", 0); reorder_lvl = item.get("reorder_level", 0)
        stock_color_tag = "low_stock" if stock <= reorder_lvl else "ok_stock"
        inventory_tree.insert("", tk.END, values=(
            idx, item["name"], item.get("category", "N/A"), stock, reorder_lvl,
            format_currency(item.get("cost_price", 0)),
            format_currency(item.get("sale_price", 0))
        ), tags=(tag, stock_color_tag), iid=item['name_key'])

def add_new_product():
    ProductEditDialog(root)

def edit_selected_product():
    if not inventory_tree: return
    sel = inventory_tree.focus()
    if not sel:
        messagebox.showwarning("Select", "Select a product to edit."); return
    item_key = sel
    if item_key in inventory:
        ProductEditDialog(root, product_data=inventory[item_key])
    else:
        messagebox.showerror("Error", "Could not find product data.")

def quick_adjust_stock():
    if not inventory_tree: return
    sel = inventory_tree.focus()
    if not sel:
        messagebox.showwarning("Select", "Select a product to adjust stock."); return
    item_key = sel
    if item_key not in inventory:
         messagebox.showerror("Error", "Could not find product data."); return

    item = inventory[item_key]
    item_name, current_stock = item['name'], item['stock']
    try:
        new_stock = simpledialog.askinteger("Adjust Stock", 
                                            f"Enter NEW total stock for '{item_name}':\n(Current stock is {current_stock})",
                                            initialvalue=current_stock, minvalue=0, parent=root)
        if new_stock is None: return
        adjust_product_stock_db(item_key, new_stock)
    except Exception as e:
        messagebox.showerror("Error", f"Invalid stock amount: {e}")

# --- NEW: Delete Selected Product ---
def delete_selected_product():
    if not inventory_tree: return
    sel = inventory_tree.focus()
    if not sel:
        messagebox.showwarning("Select", "Select a product to delete."); return
    item_key = sel
    if item_key in inventory:
        delete_product_db(item_key)
    else:
        messagebox.showerror("Error", "Could not find product data.")

# ----------------------------------------------------------------------
# ------------------- PART 4: UI CONSTRUCTION --------------------------
# ----------------------------------------------------------------------

def show_frame(frame_name):
    """Hides all frames and shows the requested one."""
    if frame_name not in frames: return
    for frame in frames.values():
        frame.pack_forget()
    frames[frame_name].pack(fill="both", expand=True)
    
    # Refresh data when switching
    if frame_name == "dashboard":
        update_main_dashboard_summary()
    elif frame_name == "billing":
        refresh_table(filter_entry.get() if filter_entry else "", type_filter.get() if type_filter else "All")
        update_billing_summary()
    elif frame_name == "inventory":
        refresh_inventory_table()
    elif frame_name == "reports":
        run_sales_report() # Run with default dates
    elif frame_name == "customers":
        refresh_customer_list()

def create_summary_card(parent, title, fg, icon="💰"):
    """Helper to create a summary card."""
    c = tk.Frame(parent, bg=CARD, highlightbackground="#E6EEF8", highlightthickness=1)
    c.pack(side="left", expand=True, fill="x", padx=6, pady=4)
    icon_lbl = tk.Label(c, text=icon, bg=CARD, fg=fg, font=("Segoe UI", 24, "bold"))
    icon_lbl.pack(side="left", padx=(12, 0))
    text_frame = tk.Frame(c, bg=CARD)
    text_frame.pack(side="left", fill="x", padx=12, pady=(8, 12))
    tk.Label(text_frame, text=title, bg=CARD, fg="#374151").pack(anchor="w")
    lbl = tk.Label(text_frame, text="Rs. 0.00", bg=CARD, fg=fg, font=("Segoe UI", 13, "bold"))
    lbl.pack(anchor="w", pady=(2,0))
    return lbl

def create_dashboard_ui(parent, style):
    """Creates the main dashboard screen."""
    global lbl_inventory_value, lbl_total_profit
    frame = ttk.Frame(parent, style="TFrame")
    
    ttk.Label(frame, text="Business Manager Dashboard", style="Header.TLabel", font=DASH_HEADER_FONT).pack(pady=(40, 20))
    
    dash_summary = ttk.Frame(frame, style="TFrame")
    dash_summary.pack(fill="x", padx=20, pady=(0,12))
    
    # --- NEW: Added Total Profit card ---
    lbl_total_profit = create_summary_card(dash_summary, "Total Profit", PROFIT, "💡")
    lbl_inventory_value = create_summary_card(dash_summary, "Total Inventory Value", "#111827", "📦")
    
    ttk.Label(frame, text="Select an option to begin.", font=APP_FONT).pack(pady=(40, 40))
    
    btn_container = ttk.Frame(frame, style="TFrame")
    btn_container.pack(pady=20)
    
    style.configure("Dash.TButton", font=DASH_BTN_FONT, padding=(30, 20), background=ACCENT, foreground="white")
    style.map("Dash.TButton", background=[('active', '#1E63D0')])
    style.configure("Dash2.TButton", font=DASH_BTN_FONT, padding=(30, 20), background=ACCENT2, foreground="white")
    style.map("Dash2.TButton", background=[('active', '#D95A5A')])
    # --- NEW: Styles for new buttons ---
    style.configure("Dash3.TButton", font=DASH_BTN_FONT, padding=(30, 20), background=SUCCESS, foreground="white")
    style.map("Dash3.TButton", background=[('active', '#218838')])
    style.configure("Dash4.TButton", font=DASH_BTN_FONT, padding=(30, 20), background=PROFIT, foreground="white")
    style.map("Dash4.TButton", background=[('active', '#138496')])

    # --- NEW: Added Reports and Customers buttons ---
    btn_row1 = ttk.Frame(btn_container, style="TFrame")
    btn_row1.pack()
    ttk.Button(btn_row1, text="Manage Billing 🧾", style="Dash.TButton", command=lambda: show_frame("billing")).pack(side="left", padx=15, pady=15)
    ttk.Button(btn_row1, text="Manage Inventory 📦", style="Dash2.TButton", command=lambda: show_frame("inventory")).pack(side="left", padx=15, pady=15)
    
    btn_row2 = ttk.Frame(btn_container, style="TFrame")
    btn_row2.pack()
    ttk.Button(btn_row2, text="View Reports 📊", style="Dash3.TButton", command=lambda: show_frame("reports")).pack(side="left", padx=15, pady=15)
    ttk.Button(btn_row2, text="Customers 👥", style="Dash4.TButton", command=lambda: show_frame("customers")).pack(side="left", padx=15, pady=15)

    return frame

def create_billing_ui(parent, style):
    """Creates the entire billing UI inside the 'parent' frame."""
    global tree, items_tree, filter_entry, type_filter, status_lbl
    global customer_entry, item_entry, qty_entry, price_entry, mode_entry
    global lbl_total_sales, lbl_total_purchases, lbl_net, items_count_lbl, type_var
    global lbl_today_sales, lbl_month_sales
    
    billing_frame = ttk.Frame(parent, style="TFrame")
    
    header = ttk.Frame(billing_frame, style="TFrame")
    header.pack(fill="x", padx=20, pady=14)
    ttk.Button(header, text="⬅ Dashboard", style="TButton", command=lambda: show_frame("dashboard")).pack(side="left", padx=(0, 20))
    ttk.Label(header, text="Billing & Transactions", style="Header.TLabel").pack(side="left")
    action_frame = ttk.Frame(header, style="TFrame"); action_frame.pack(side="right")
    make_icon_btn(action_frame, "🧾 Invoice PDF", create_invoice_pdf, "#2E8B57")
    make_icon_btn(action_frame, "🏢 Business Info", edit_business_profile, "#6C63FF")
    
    # --- Dashboard Cards ---
    dash = ttk.Frame(billing_frame, style="TFrame")
    dash.pack(fill="x", padx=20, pady=(0,12))
    
    # --- NEW: Added Today and Month cards ---
    lbl_today_sales = create_summary_card(dash, "Today's Sales", SUCCESS, "☀️")
    lbl_month_sales = create_summary_card(dash, "This Month's Sales", SUCCESS, "📅")
    lbl_total_sales = create_summary_card(dash, "Total Sales", ACCENT, "📈")
    lbl_total_purchases = create_summary_card(dash, "Total Purchases", ACCENT2, "📉")
    lbl_net = create_summary_card(dash, "Net (Sales - Purchases)", "#111827", "💰")
    
    content = ttk.Frame(billing_frame, style="TFrame")
    content.pack(fill="both", expand=True, padx=20, pady=10)
    
    # --- Left Panel (Form) ---
    left = ttk.Frame(content, style="Card.TFrame")
    left.pack(side="left", fill="y", padx=(0,12), ipadx=8, ipady=8)
    
    tk.Label(left, text="Add / Edit Bill", bg=CARD, fg="#111827", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(4,8))
    type_var = tk.StringVar(value="Sale")
    tf = ttk.Frame(left, style="Card.TFrame"); tf.pack(anchor="w")
    tk.Radiobutton(tf, text="Sale", variable=type_var, value="Sale", bg=CARD, bd=0, activebackground=CARD).pack(side="left")
    tk.Radiobutton(tf, text="Purchase", variable=type_var, value="Purchase", bg=CARD, bd=0, activebackground=CARD).pack(side="left", padx=8)

    customer_entry = labeled_entry(left, "Customer / Supplier")
    item_entry = labeled_entry(left, "Item")
    item_entry.bind("<Return>", lambda e: (auto_fill_item_price(), qty_entry.focus()))
    
    row = ttk.Frame(left, style="Card.TFrame"); row.pack(fill="x", pady=(8,0))
    tk.Label(row, text="Quantity", bg=CARD).grid(row=0, column=0, sticky="w")
    tk.Label(row, text="Price", bg=CARD).grid(row=0, column=1, sticky="w", padx=(12,0))
    
    int_vcmd = (root.register(validate_integer), '%P')
    float_vcmd = (root.register(validate_float), '%P')
    qty_entry = ttk.Entry(row, width=12, validate="key", validatecommand=int_vcmd)
    qty_entry.grid(row=1, column=0, sticky="w", pady=(2,0))
    qty_entry.bind("<Return>", lambda e: price_entry.focus())
    price_entry = ttk.Entry(row, width=12, validate="key", validatecommand=float_vcmd)
    price_entry.grid(row=1, column=1, sticky="w", padx=(12,0), pady=(2,0))
    price_entry.bind("<Return>", lambda e: add_item_to_current())
    
    mode_entry = labeled_entry(left, "Mode (Cash/Credit)")
    mode_entry.bind("<Return>", lambda e: add_bill())

    tk.Label(left, text="Items (for multi-item invoice)", bg=CARD, fg="#374151").pack(anchor="w", pady=(8,2))
    items_container = tk.Frame(left, bg=CARD); items_container.pack(fill="x", expand=False)
    items_tree = ttk.Treeview(items_container, columns=("SNo", "Name", "Qty", "Price", "Total"), show="headings", height=5)
    for c, w, anchor in (("SNo", 40, "center"), ("Name", 120, "w"), ("Qty", 50, "center"), ("Price", 80, "e"), ("Total", 80, "e")):
        items_tree.heading(c, text=c); items_tree.column(c, width=w, anchor=anchor)
    items_tree.pack(side="left", fill="x", expand=True)
    items_vsb = ttk.Scrollbar(items_container, orient="vertical", command=items_tree.yview)
    items_tree.configure(yscroll=items_vsb.set); items_vsb.pack(side="right", fill="y")
    items_count_lbl = tk.Label(left, text="Items: 0", bg=CARD); items_count_lbl.pack(anchor="w", pady=(4,0))

    items_btn_frame = tk.Frame(left, bg=CARD); items_btn_frame.pack(fill="x", pady=(6,0))
    make_small_btn(items_btn_frame, "➕ Add", add_item_to_current)
    make_small_btn(items_btn_frame, "🗑️ Remove", remove_selected_item)
    make_small_btn(items_btn_frame, "♻️ Clear List", clear_current_items)

    btn_frame = tk.Frame(left, bg=CARD); btn_frame.pack(fill="x", pady=12)
    make_btn(btn_frame, "➕ Add Bill", add_bill, ACCENT, style)
    make_btn(btn_frame, "✏️ Edit Bill", edit_bill, "#0EA5A4", style)
    make_btn(btn_frame, "🗑️ Delete", delete_bill, ACCENT2, style)
    make_btn(btn_frame, "♻️ Clear", clear_entries, "#6B7280", style)
    
    ttk.Separator(left).pack(fill="x", pady=10)
    make_btn(left, "📊 Show Ledger (by name)", show_ledger, "#374151", style)
    tk.Label(left, text="", bg=CARD).pack(pady=4)
    # --- NEW: Renamed export function ---
    make_btn(left, "📤 Export All (Excel)", export_bills_excel, ACCENT, style)

    # --- Right Panel (Table) ---
    right = ttk.Frame(content, style="Card.TFrame")
    right.pack(side="left", fill="both", expand=True, ipadx=8, ipady=8)

    top_right = ttk.Frame(right, style="Card.TFrame"); top_right.pack(fill="x")
    tk.Label(top_right, text="Search / Filter", bg=CARD).pack(side="left")
    filter_entry = ttk.Entry(top_right, width=30); filter_entry.pack(side="left", padx=8)
    filter_entry.bind("<KeyRelease>", on_search_change)
    type_filter = tk.StringVar(value="All")
    ttk.OptionMenu(top_right, type_filter, "All", "All", "Sale", "Purchase", command=on_filter_change).pack(side="left")

    table_container = ttk.Frame(right); table_container.pack(fill="both", expand=True, pady=(12,0))
    # --- NEW: Added Date column to table ---
    columns = ("S.No", "BillNo", "Date", "Type", "Customer", "Item", "Qty", "Total", "Mode")
    tree = ttk.Treeview(table_container, columns=columns, show="headings", selectmode="browse", height=18)
    col_config = [
        ("S.No", 40, "center"), ("BillNo", 50, "center"), ("Date", 90, "center"),
        ("Type", 70, "center"), ("Customer", 150, "w"), ("Item", 150, "w"),
        ("Qty", 50, "center"), ("Total", 100, "e"), ("Mode", 70, "center")
    ]
    for col, w, a in col_config:
        tree.heading(col, text=col); tree.column(col, width=w, anchor=a)
        
    tree.tag_configure("odd", background=ROW_ODD); tree.tag_configure("even", background=ROW_EVEN)
    vsb = ttk.Scrollbar(table_container, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(table_container, orient="horizontal", command=tree.xview)
    tree.configure(yscroll=vsb.set, xscroll=hsb.set)
    vsb.pack(side="right", fill="y"); hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)
    tree.bind("<<TreeviewSelect>>", on_row_select)
    
    return billing_frame

def create_inventory_ui(parent, style):
    """Creates the new inventory management UI."""
    global inventory_tree
    
    frame = ttk.Frame(parent, style="TFrame")
    header = ttk.Frame(frame, style="TFrame"); header.pack(fill="x", padx=20, pady=14)
    ttk.Button(header, text="⬅ Dashboard", style="TButton", command=lambda: show_frame("dashboard")).pack(side="left", padx=(0, 20))
    ttk.Label(header, text="Inventory Management", style="Header.TLabel").pack(side="left")
    
    content = ttk.Frame(frame, style="Card.TFrame")
    content.pack(fill="both", expand=True, padx=20, pady=10, ipadx=10, ipady=10)
    
    btn_frame = ttk.Frame(content, style="Card.TFrame"); btn_frame.pack(fill="x", pady=(5, 15))
    make_btn(btn_frame, "➕ Add New Product", add_new_product, ACCENT, style)
    make_btn(btn_frame, "✏️ Edit Selected", edit_selected_product, "#0EA5A4", style)
    # --- NEW: Delete button ---
    make_btn(btn_frame, "🗑️ Delete Selected", delete_selected_product, ACCENT2, style)
    make_btn(btn_frame, "🔄 Quick Adjust Stock", quick_adjust_stock, "#EAB308", style, fg="#111827")
    
    btn_frame2 = ttk.Frame(content, style="Card.TFrame"); btn_frame2.pack(fill="x", pady=(0, 15))
    make_btn(btn_frame2, "📊 View Low Stock", lambda: refresh_inventory_table(low_stock_only=True), ACCENT2, style)
    make_btn(btn_frame2, "📋 Show All Stock", lambda: refresh_inventory_table(low_stock_only=False), "#6B7280", style)
    # --- NEW: Export Button ---
    make_btn(btn_frame2, "📤 Export (Excel)", export_inventory_excel, SUCCESS, style)

    table_container = ttk.Frame(content); table_container.pack(fill="both", expand=True, pady=(12,0))
    columns = ("S.No", "Product Name", "Category", "Stock", "Reorder Lvl", "Cost Price", "Sale Price")
    inventory_tree = ttk.Treeview(table_container, columns=columns, show="headings", selectmode="browse", height=20)
    
    col_config = [
        ("S.No", 60, "center"), ("Product Name", 300, "w"), ("Category", 150, "w"),
        ("Stock", 100, "center"), ("Reorder Lvl", 100, "center"),
        ("Cost Price", 120, "e"), ("Sale Price", 120, "e")
    ]
    for col, w, a in col_config:
        inventory_tree.heading(col, text=col); inventory_tree.column(col, width=w, anchor=a)
        
    inventory_tree.tag_configure("odd", background=ROW_ODD); inventory_tree.tag_configure("even", background=ROW_EVEN)
    inventory_tree.tag_configure("low_stock", foreground=ACCENT2, font=(APP_FONT[0], 10, "bold"))
    inventory_tree.tag_configure("ok_stock", foreground=TEXT)
    
    vsb = ttk.Scrollbar(table_container, orient="vertical", command=inventory_tree.yview)
    inventory_tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y")
    inventory_tree.pack(fill="both", expand=True)
    inventory_tree.bind("<Double-1>", lambda e: edit_selected_product())

    return frame

# --- NEW: Reports Tab ---
def create_reports_ui(parent, style):
    """Creates the new Reports UI."""
    global report_tree, report_start_date_entry, report_end_date_entry
    
    frame = ttk.Frame(parent, style="TFrame")
    header = ttk.Frame(frame, style="TFrame"); header.pack(fill="x", padx=20, pady=14)
    ttk.Button(header, text="⬅ Dashboard", style="TButton", command=lambda: show_frame("dashboard")).pack(side="left", padx=(0, 20))
    ttk.Label(header, text="Sales & Profit Reports", style="Header.TLabel").pack(side="left")
    
    content = ttk.Frame(frame, style="Card.TFrame")
    content.pack(fill="both", expand=True, padx=20, pady=10, ipadx=10, ipady=10)
    
    # --- Date Filter Bar ---
    filter_frame = ttk.Frame(content, style="Card.TFrame")
    filter_frame.pack(fill="x", pady=(5, 10))
    
    ttk.Label(filter_frame, text="Start Date:", background=CARD).pack(side="left", padx=(10, 5))
    
    # Use tkcalendar DateEntry if available, otherwise fall back to simple Entry
    if CALENDAR_ENABLED:
        report_start_date_entry = DateEntry(filter_frame, width=12, background=ACCENT, foreground='white', borderwidth=2,
                                            date_pattern='y-mm-dd')
    else:
        report_start_date_entry = ttk.Entry(filter_frame, width=14)
        report_start_date_entry.insert(0, (datetime.date.today() - datetime.timedelta(days=30)).strftime('%Y-%m-%d'))
    report_start_date_entry.pack(side="left", padx=5)

    ttk.Label(filter_frame, text="End Date:", background=CARD).pack(side="left", padx=(10, 5))
    if CALENDAR_ENABLED:
        report_end_date_entry = DateEntry(filter_frame, width=12, background=ACCENT, foreground='white', borderwidth=2,
                                          date_pattern='y-mm-dd')
    else:
        report_end_date_entry = ttk.Entry(filter_frame, width=14)
        report_end_date_entry.insert(0, datetime.date.today().strftime('%Y-%m-%d'))
    report_end_date_entry.pack(side="left", padx=5)

    make_btn(filter_frame, "📊 Run Report", run_sales_report, SUCCESS, style)
    # --- NEW: Export Button ---
    make_btn(filter_frame, "📤 Export (Excel)", export_report_excel, PROFIT, style)
    
    # --- Report Table ---
    table_container = ttk.Frame(content); table_container.pack(fill="both", expand=True, pady=(12,0))
    columns = ("Item Name", "Units Sold", "Total Revenue", "Total Cost", "Total Profit")
    report_tree = ttk.Treeview(table_container, columns=columns, show="headings", selectmode="browse", height=20)
    
    col_config = [
        ("Item Name", 300, "w"), ("Units Sold", 100, "center"),
        ("Total Revenue", 150, "e"), ("Total Cost", 150, "e"), ("Total Profit", 150, "e")
    ]
    for col, w, a in col_config:
        report_tree.heading(col, text=col); report_tree.column(col, width=w, anchor=a)
        
    report_tree.tag_configure("odd", background=ROW_ODD); report_tree.tag_configure("even", background=ROW_EVEN)
    vsb = ttk.Scrollbar(table_container, orient="vertical", command=report_tree.yview)
    report_tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y")
    report_tree.pack(fill="both", expand=True)

    return frame

def run_sales_report():
    """Queries DB and populates the sales report tree."""
    if not report_tree: return
    
    try:
        start_date = report_start_date_entry.get()
        end_date = report_end_date_entry.get()
        datetime.datetime.strptime(start_date, '%Y-%m-%d')
        datetime.datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.")
        return

    for i in report_tree.get_children():
        report_tree.delete(i)
        
    conn = db_connect()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT 
        name, 
        SUM(qty) as TotalUnits,
        SUM(total) as TotalRevenue,
        SUM(cost_price * qty) as TotalCost,
        SUM(total) - SUM(cost_price * qty) as TotalProfit
    FROM bill_items
    WHERE bill_id IN (
        SELECT id FROM bills WHERE type = 'Sale' AND date BETWEEN ? AND ?
    )
    GROUP BY name
    ORDER BY TotalProfit DESC
    """, (start_date, end_date))
    
    rows = cursor.fetchall()
    conn.close()
    
    total_profit_summary = 0
    total_revenue_summary = 0
    
    for idx, row in enumerate(rows):
        tag = "even" if idx % 2 == 0 else "odd"
        report_tree.insert("", tk.END, values=(
            row['name'],
            row['TotalUnits'],
            format_currency(row['TotalRevenue']),
            format_currency(row['TotalCost']),
            format_currency(row['TotalProfit'])
        ), tags=(tag,))
        total_revenue_summary += row['TotalRevenue']
        total_profit_summary += row['TotalProfit']

    # Add a summary footer
    report_tree.insert("", tk.END, values=(
        "--- TOTAL ---", "",
        format_currency(total_revenue_summary),
        "",
        format_currency(total_profit_summary)
    ), tags=("total_row",))
    report_tree.tag_configure("total_row", font=("Segoe UI", 11, "bold"), background="#EAECEE")
    
    set_status(f"Report generated for {start_date} to {end_date}")

# --- NEW: Customers Tab ---
def create_customers_ui(parent, style):
    """Creates the new Customers UI."""
    global customer_tree
    
    frame = ttk.Frame(parent, style="TFrame")
    header = ttk.Frame(frame, style="TFrame"); header.pack(fill="x", padx=20, pady=14)
    ttk.Button(header, text="⬅ Dashboard", style="TButton", command=lambda: show_frame("dashboard")).pack(side="left", padx=(0, 20))
    ttk.Label(header, text="Customer List", style="Header.TLabel").pack(side="left")
    
    content = ttk.Frame(frame, style="Card.TFrame")
    content.pack(fill="both", expand=True, padx=20, pady=10, ipadx=10, ipady=10)

    btn_frame = ttk.Frame(content, style="Card.TFrame"); btn_frame.pack(fill="x", pady=(5, 10))
    make_btn(btn_frame, "🔄 Refresh List", refresh_customer_list, SUCCESS, style)
    # --- NEW: Export Button ---
    make_btn(btn_frame, "📤 Export (Excel)", export_customers_excel, PROFIT, style)
    ttk.Label(btn_frame, text=" (This is a read-only view of all customers from your bills)", background=CARD).pack(side="left", padx=10)

    table_container = ttk.Frame(content); table_container.pack(fill="both", expand=True, pady=(12,0))
    columns = ("Customer Name", "Total Bills", "Total Spent")
    customer_tree = ttk.Treeview(table_container, columns=columns, show="headings", selectmode="browse", height=20)
    
    col_config = [
        ("Customer Name", 300, "w"), ("Total Bills", 150, "center"), ("Total Spent", 200, "e")
    ]
    for col, w, a in col_config:
        customer_tree.heading(col, text=col); customer_tree.column(col, width=w, anchor=a)
        
    customer_tree.tag_configure("odd", background=ROW_ODD); customer_tree.tag_configure("even", background=ROW_EVEN)
    vsb = ttk.Scrollbar(table_container, orient="vertical", command=customer_tree.yview)
    customer_tree.configure(yscroll=vsb.set); vsb.pack(side="right", fill="y")
    customer_tree.pack(fill="both", expand=True)

    return frame

def refresh_customer_list():
    """Queries DB and populates the customer list."""
    if not customer_tree: return
    for i in customer_tree.get_children():
        customer_tree.delete(i)
        
    conn = db_connect()
    cursor = conn.cursor()
    # Query for customers from Sales only
    cursor.execute("""
    SELECT 
        customer, 
        COUNT(id) as TotalBills,
        SUM(grand_total) as TotalSpent
    FROM bills
    WHERE type = 'Sale'
    GROUP BY customer
    ORDER BY TotalSpent DESC
    """)
    
    rows = cursor.fetchall()
    conn.close()
    
    for idx, row in enumerate(rows):
        tag = "even" if idx % 2 == 0 else "odd"
        customer_tree.insert("", tk.END, values=(
            row['customer'],
            row['TotalBills'],
            format_currency(row['TotalSpent'])
        ), tags=(tag,))
    
    set_status(f"Loaded {len(rows)} customers")

# --- UI: Helper functions for creating widgets ---
def labeled_entry(parent, label):
    tk.Label(parent, text=label, bg=CARD, fg="#374151").pack(anchor="w", pady=(8,2))
    e = ttk.Entry(parent, width=28)
    e.pack(fill="x")
    return e

def make_icon_btn(parent, text, command, bg_color):
    b = tk.Button(parent, text=text, command=command, bg=bg_color, fg="white",
                  bd=0, activebackground=bg_color, font=("Segoe UI", 10, "bold"), padx=8, pady=4)
    b.pack(side="right", padx=6)
    return b

def make_small_btn(parent, text, cmd):
    b = ttk.Button(parent, text=text, command=cmd, style="TButton")
    b.pack(side="left", padx=4)
    return b

def make_btn(parent, text, cmd, bg, style, fg="white"):
    style_name = f"{''.join(filter(str.isalnum, text))}.TButton"
    style.configure(style_name, background=bg, foreground=fg, font=("Segoe UI", 10, "bold"), padding=(10, 6))
    style.map(style_name, background=[('active', bg)])
    b = ttk.Button(parent, text=text, command=cmd, style=style_name)
    b.pack(side="left", padx=6)
    return b

# ----------------------------------------------------------------------
# ------------------- PART 5: MAIN EXECUTION ---------------------------
# ----------------------------------------------------------------------

def main():
    global root, status_lbl, frames
    
    init_db()
    
    root = tk.Tk()
    root.title("Business Transaction Manager")
    root.geometry("1200x720")
    root.configure(bg=BG)
    root.option_add("*Font", APP_FONT)
    
    style = ttk.Style(root)
    style.theme_use("clam")
    style.configure("TFrame", background=BG)
    style.configure("Card.TFrame", background=CARD)
    style.configure("TLabel", background=BG, foreground=TEXT)
    style.configure("Header.TLabel", font=HEADER_FONT, background=BG, foreground=TEXT)
    style.configure("Treeview", background=ROW_ODD, fieldbackground=ROW_ODD, rowheight=26, font=APP_FONT)
    style.configure("Treeview.Heading", font=(APP_FONT[0], 10, "bold"))
    style.configure("TButton", padding=6, font=("Segoe UI", 10, "bold"))
    style.configure("Accent.TButton", background=ACCENT, foreground="white")
    style.map("Accent.TButton", background=[('active', '#1E63D0')])
    
    main_container = ttk.Frame(root, style="TFrame")
    main_container.pack(fill="both", expand=True)
    
    # --- Create all frames ---
    frames["dashboard"] = create_dashboard_ui(main_container, style)
    frames["billing"] = create_billing_ui(main_container, style)
    frames["inventory"] = create_inventory_ui(main_container, style)
    frames["reports"] = create_reports_ui(main_container, style)
    frames["customers"] = create_customers_ui(main_container, style)
    
    status_frame = ttk.Frame(root, style="TFrame")
    status_frame.pack(fill="x", padx=20, pady=(8,12))
    status_lbl = tk.Label(status_frame, text="Loading...", bg=BG, fg="#475569")
    status_lbl.pack(side="left")

    load_data()
    
    show_frame("dashboard")
    
    set_status("Welcome — Business Manager ready", timeout=2500)
    root.mainloop()

if __name__ == "__main__":
    main()
